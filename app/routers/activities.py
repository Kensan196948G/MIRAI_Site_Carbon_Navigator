import io

from fastapi import APIRouter, Depends, File, HTTPException, UploadFile
from fastapi.responses import Response
from openpyxl import load_workbook
from sqlalchemy.orm import Session

from .. import crud, schemas
from ..database import get_db
from ..security import get_current_user, require_at_least
from ..services.reporter import generate_activity_import_template

router = APIRouter(prefix="/api/activities", tags=["activities"])


@router.post("", response_model=schemas.ActivityDataRead, status_code=201)
def create_activity(
    activity: schemas.ActivityDataCreate,
    db: Session = Depends(get_db),
    user=Depends(require_at_least("site")),
):
    project = crud.get_project(db, activity.project_id)
    if not project:
        raise HTTPException(status_code=404, detail="Project not found")
    if user.role == "site" and user.branch and project.branch != user.branch:
        raise HTTPException(status_code=403, detail="他支店の工事へは登録できません")
    if crud.is_month_closed(db, activity.project_id, activity.target_month):
        raise HTTPException(status_code=400, detail="対象月は締め済みのため登録できません")
    try:
        return crud.create_activity_data(db, activity, actor=user.username)
    except ValueError as e:
        raise HTTPException(status_code=400, detail=str(e)) from e


@router.get("", response_model=list[schemas.ActivityDataRead])
def list_activities(
    project_id: str | None = None,
    target_month: str | None = None,
    db: Session = Depends(get_db),
    user=Depends(get_current_user),
):
    if project_id and not crud.has_project_access(db, user, project_id):
        raise HTTPException(status_code=403, detail="Project access denied")
    return crud.list_activity_data(db, project_id=project_id, target_month=target_month)


@router.get("/template")
def download_template(user=Depends(require_at_least("site"))):
    filename = "activity_import_template.xlsx"
    return Response(
        content=generate_activity_import_template(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@router.post("/import", response_model=schemas.ImportResult, status_code=200)
async def import_activities(
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
    user=Depends(require_at_least("site")),
):
    if not file.filename or not file.filename.lower().endswith((".xlsx", ".xlsm")):
        raise HTTPException(status_code=400, detail="Only .xlsx files are supported")
    try:
        raw = await file.read()
        wb = load_workbook(io.BytesIO(raw), data_only=True)
    except Exception:
        raise HTTPException(status_code=400, detail="Could not parse Excel file") from None
    ws = wb.active
    imported = 0
    skipped = 0
    errors: list[str] = []
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        raise HTTPException(status_code=400, detail="Excel file is empty")
    headers = [str(h).strip().lower() if h else "" for h in rows[0]]
    required = {"project_id", "target_month", "category", "item_name", "quantity", "unit"}
    if not required.issubset(set(headers)):
        raise HTTPException(
            status_code=400,
            detail=f"Required columns missing: {', '.join(sorted(required - set(headers)))}",
        )
    for idx, row in enumerate(rows[1:], start=2):
        record = dict(zip(headers, row, strict=False))
        if not any(record.values()):
            continue
        payload = {
            "project_id": str(record.get("project_id") or "").strip(),
            "target_month": str(record.get("target_month") or "").strip(),
            "category": str(record.get("category") or "").strip(),
            "item_name": str(record.get("item_name") or "").strip(),
            "quantity": record.get("quantity"),
            "unit": str(record.get("unit") or "").strip(),
            "source_file": str(record.get("source_file") or "").strip() or None,
            "note": str(record.get("note") or "").strip() or None,
        }
        try:
            schema = schemas.ActivityDataCreate(**payload)
        except Exception as e:
            errors.append(f"Row {idx}: {e}")
            skipped += 1
            continue
        if not crud.get_project(db, schema.project_id):
            errors.append(f"Row {idx}: Project {schema.project_id} not found")
            skipped += 1
            continue
        if crud.is_month_closed(db, schema.project_id, schema.target_month):
            errors.append(f"Row {idx}: 対象月 {schema.target_month} は締め済み")
            skipped += 1
            continue
        try:
            crud.create_activity_data(db, schema, actor=user.username)
            imported += 1
        except ValueError as e:
            errors.append(f"Row {idx}: {e}")
            skipped += 1
    return schemas.ImportResult(imported=imported, skipped=skipped, errors=errors)


@router.put("/{activity_id}/approve", response_model=schemas.ActivityDataRead)
def approve_activity(
    activity_id: str,
    body: schemas.ActivityDataApprove = schemas.ActivityDataApprove(),
    db: Session = Depends(get_db),
    user=Depends(require_at_least("reviewer")),
):
    activity = crud.get_activity(db, activity_id)
    if not activity:
        raise HTTPException(status_code=404, detail="Activity not found")
    if crud.is_month_closed(db, activity.project_id, activity.target_month):
        raise HTTPException(status_code=400, detail="対象月は締め済みのため承認を変更できません")
    activity = crud.approve_activity(db, activity_id, body.approved, user.username)
    return activity


@router.put("/{activity_id}/approval", response_model=schemas.ActivityDataRead)
def approval_action(
    activity_id: str,
    body: schemas.ApprovalAction,
    db: Session = Depends(get_db),
    user=Depends(require_at_least("site")),
):
    activity = crud.get_activity(db, activity_id)
    if not activity:
        raise HTTPException(status_code=404, detail="Activity not found")
    if crud.is_month_closed(db, activity.project_id, activity.target_month):
        raise HTTPException(status_code=400, detail="対象月は締め済みです")
    role_needed = {
        "submit": "site",
        "withdraw": "site",
        "approve_branch": "reviewer",
        "reject": "reviewer",
        "approve_env": "reviewer",
    }.get(body.action)
    if role_needed and not (
        user.role == "admin"
        or (role_needed == "reviewer" and user.role in {"reviewer", "admin"})
        or (role_needed == "site" and user.role in {"site", "reviewer", "admin"})
    ):
        raise HTTPException(status_code=403, detail="Insufficient permissions")
    try:
        return crud.transition_approval(
            db, activity_id, body.action, user.username, body.comment
        )
    except ValueError as e:
        raise HTTPException(status_code=400, detail=str(e)) from e


@router.put("/{activity_id}", response_model=schemas.ActivityDataRead)
def update_activity(
    activity_id: str,
    body: schemas.ActivityDataUpdate,
    db: Session = Depends(get_db),
    user=Depends(require_at_least("site")),
):
    existing = crud.get_activity(db, activity_id)
    if not existing:
        raise HTTPException(status_code=404, detail="Activity not found")
    if crud.is_month_closed(db, existing.project_id, existing.target_month):
        raise HTTPException(status_code=400, detail="対象月は締め済みのため更新できません")
    activity = crud.update_activity(db, activity_id, body, user.username)
    return activity


@router.delete("/{activity_id}", status_code=204)
def delete_activity(
    activity_id: str,
    db: Session = Depends(get_db),
    user=Depends(require_at_least("site")),
):
    existing = crud.get_activity(db, activity_id)
    if not existing:
        raise HTTPException(status_code=404, detail="Activity not found")
    if crud.is_month_closed(db, existing.project_id, existing.target_month):
        raise HTTPException(status_code=400, detail="対象月は締め済みのため削除できません")
    if not crud.delete_activity(db, activity_id, user.username):
        raise HTTPException(status_code=404, detail="Activity not found")


@router.post("/bulk", response_model=list[schemas.ActivityDataRead], status_code=201)
def bulk_create_activities(
    activities: list[schemas.ActivityDataCreate],
    db: Session = Depends(get_db),
    user=Depends(require_at_least("site")),
):
    results = []
    errors = []
    for i, activity in enumerate(activities):
        project = crud.get_project(db, activity.project_id)
        if not project:
            errors.append(f"Row {i}: Project {activity.project_id} not found")
            continue
        if crud.is_month_closed(db, activity.project_id, activity.target_month):
            errors.append(f"Row {i}: 対象月 {activity.target_month} は締め済み")
            continue
        try:
            result = crud.create_activity_data(db, activity, actor=user.username)
            results.append(result)
        except ValueError as e:
            errors.append(f"Row {i}: {str(e)}")
    if errors and not results:
        raise HTTPException(status_code=400, detail="; ".join(errors))
    return results


@router.post("/copy-previous", response_model=schemas.CopyPreviousResult)
def copy_previous(
    body: schemas.CopyPreviousRequest,
    db: Session = Depends(get_db),
    user=Depends(require_at_least("site")),
):
    if not crud.get_project(db, body.project_id):
        raise HTTPException(status_code=404, detail="Project not found")
    if crud.is_month_closed(db, body.project_id, body.to_month):
        raise HTTPException(status_code=400, detail="コピー先の対象月は締め済みです")
    return crud.copy_previous_month(
        db, body.project_id, body.from_month, body.to_month, user.username
    )


@router.get("/{activity_id}/comments", response_model=list[schemas.ActivityCommentRead])
def list_comments(
    activity_id: str,
    db: Session = Depends(get_db),
    user=Depends(get_current_user),
):
    activity = crud.get_activity(db, activity_id)
    if not activity:
        raise HTTPException(status_code=404, detail="Activity not found")
    return crud.list_activity_comments(db, activity_id)


@router.get("/{activity_id}/history", response_model=list[schemas.ActivityChangeRead])
def list_activity_history(
    activity_id: str,
    db: Session = Depends(get_db),
    user=Depends(get_current_user),
):
    activity = crud.get_activity(db, activity_id)
    if not activity:
        raise HTTPException(status_code=404, detail="Activity not found")
    return crud.list_activity_changes(db, activity_id)


@router.post("/{activity_id}/comments", response_model=schemas.ActivityCommentRead, status_code=201)
def add_comment(
    activity_id: str,
    body: schemas.ActivityCommentCreate,
    db: Session = Depends(get_db),
    user=Depends(require_at_least("site")),
):
    activity = crud.get_activity(db, activity_id)
    if not activity:
        raise HTTPException(status_code=404, detail="Activity not found")
    return crud.add_activity_comment(db, activity_id, body.content.strip(), user.username)


@router.delete("/{activity_id}/comments/{comment_id}", status_code=204)
def delete_comment(
    activity_id: str,
    comment_id: str,
    db: Session = Depends(get_db),
    user=Depends(require_at_least("reviewer")),
):
    if not crud.delete_activity_comment(db, comment_id, user.username):
        raise HTTPException(status_code=404, detail="Comment not found")
