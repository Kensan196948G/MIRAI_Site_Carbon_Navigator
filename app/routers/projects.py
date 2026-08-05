import io

from fastapi import APIRouter, Depends, File, HTTPException, UploadFile
from fastapi.responses import Response
from openpyxl import load_workbook
from sqlalchemy.orm import Session
from .. import crud, schemas
from ..database import get_db
from ..security import get_current_user, require_at_least
from ..services.reporter import generate_project_import_template

router = APIRouter(prefix="/api/projects", tags=["projects"])


@router.get("/template")
def download_template(user=Depends(require_at_least("site"))):
    filename = "project_import_template.xlsx"
    return Response(
        content=generate_project_import_template(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@router.post("/import", response_model=schemas.ProjectImportResult)
async def import_projects(
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
    user=Depends(require_at_least("site")),
):
    if not file.filename or not file.filename.lower().endswith((".xlsx", ".xlsm")):
        raise HTTPException(status_code=400, detail="Only .xlsx files are supported")
    try:
        wb = load_workbook(io.BytesIO(await file.read()), data_only=True)
    except Exception:
        raise HTTPException(status_code=400, detail="Could not parse Excel file")
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        raise HTTPException(status_code=400, detail="Excel file is empty")
    headers = [str(h).strip().lower() if h else "" for h in rows[0]]
    required = {"name", "branch", "work_type", "start_date", "end_date"}
    if not required.issubset(set(headers)):
        raise HTTPException(status_code=400, detail="Required columns missing")
    imported = 0
    skipped = 0
    errors: list[str] = []
    for idx, row in enumerate(rows[1:], start=2):
        record = dict(zip(headers, row))
        if not any(record.values()):
            continue
        try:
            start_date = record.get("start_date")
            end_date = record.get("end_date")
            if hasattr(start_date, "strftime"):
                start_date = start_date.strftime("%Y-%m-%d")
            if hasattr(end_date, "strftime"):
                end_date = end_date.strftime("%Y-%m-%d")
            schema = schemas.ProjectCreate(
                name=str(record.get("name") or "").strip(),
                branch=str(record.get("branch") or "").strip(),
                work_type=str(record.get("work_type") or "").strip(),
                start_date=__import__("datetime").date.fromisoformat(str(start_date).strip()),
                end_date=__import__("datetime").date.fromisoformat(str(end_date).strip()),
                description=str(record.get("description") or "").strip() or None,
                close_day=int(record["close_day"]) if record.get("close_day") not in (None, "") else None,
                created_by=user.username,
            )
        except Exception as e:
            errors.append(f"Row {idx}: {e}")
            skipped += 1
            continue
        try:
            crud.create_project(db, schema, actor=user.username)
            imported += 1
        except Exception as e:
            errors.append(f"Row {idx}: {e}")
            skipped += 1
    return schemas.ProjectImportResult(imported=imported, skipped=skipped, errors=errors)


@router.post("", response_model=schemas.ProjectRead, status_code=201)
def create_project(
    project: schemas.ProjectCreate,
    db: Session = Depends(get_db),
    user=Depends(require_at_least("site")),
):
    if user.role == "site" and user.branch:
        project.branch = user.branch
    return crud.create_project(db, project, actor=user.username)


@router.get("", response_model=list[schemas.ProjectRead])
def list_projects(
    db: Session = Depends(get_db),
    user=Depends(get_current_user),
):
    return crud.list_projects_for_user(db, user)


@router.get("/{project_id}", response_model=schemas.ProjectRead)
def get_project(
    project_id: str,
    db: Session = Depends(get_db),
    user=Depends(get_current_user),
):
    project = crud.get_project(db, project_id)
    if not project:
        raise HTTPException(status_code=404, detail="Project not found")
    if not crud.has_project_access(db, user, project_id):
        raise HTTPException(status_code=403, detail="Project access denied")
    return project


@router.put("/{project_id}", response_model=schemas.ProjectRead)
def update_project(
    project_id: str,
    body: schemas.ProjectUpdate,
    db: Session = Depends(get_db),
    user=Depends(require_at_least("site")),
):
    existing = crud.get_project(db, project_id)
    if not existing:
        raise HTTPException(status_code=404, detail="Project not found")
    if user.role == "site" and user.branch and existing.branch != user.branch:
        raise HTTPException(status_code=403, detail="他支店の工事は変更できません")
    project = crud.update_project(db, project_id, body, user.username)
    return project


@router.delete("/{project_id}", status_code=204)
def delete_project(
    project_id: str,
    db: Session = Depends(get_db),
    user=Depends(require_at_least("admin")),
):
    project = crud.get_project(db, project_id)
    if not project:
        raise HTTPException(status_code=404, detail="Project not found")
    if user.role == "site" and user.branch and project.branch != user.branch:
        raise HTTPException(status_code=403, detail="他支店の工事は削除できません")
    if not crud.delete_project(db, project_id, user.username):
        raise HTTPException(status_code=404, detail="Project not found")
