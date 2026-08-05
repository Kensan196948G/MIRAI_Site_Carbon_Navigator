import csv
import datetime
import io

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.lib.units import mm
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.cidfonts import UnicodeCIDFont
from reportlab.platypus import (
    Paragraph,
    SimpleDocTemplate,
    Spacer,
    Table,
    TableStyle,
)

from .reduction import get_reduction_suggestions
from .scope import scope_summary_from_results

try:
    pdfmetrics.registerFont(UnicodeCIDFont("HeiseiKakuGo-W5"))
    _JAPANESE_FONT = "HeiseiKakuGo-W5"
except Exception:
    _JAPANESE_FONT = "Helvetica"


def generate_monthly_report_excel(project, month: str, results_summary: list[dict]) -> bytes:
    """
    project: ProjectRead or ORM model with .name, .branch, .work_type
    month: "YYYY-MM"
    results_summary: list of dicts with keys: category, item_name, quantity, unit, factor_value, co2_kg
    Returns bytes of xlsx file.
    """
    wb = Workbook()

    # --- Sheet 1: 月次CO2レポート ---
    ws1 = wb.active
    ws1.title = "月次CO2レポート"

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="2E75B6", end_color="2E75B6", fill_type="solid")
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    # Project info section
    ws1["A1"] = "プロジェクト名"
    ws1["B1"] = getattr(project, "name", "")
    ws1["A2"] = "支店"
    ws1["B2"] = getattr(project, "branch", "")
    ws1["A3"] = "工種"
    ws1["B3"] = getattr(project, "work_type", "")
    ws1["A4"] = "対象月"
    ws1["B4"] = month
    ws1["A5"] = "レポート作成日"
    ws1["B5"] = datetime.date.today().strftime("%Y-%m-%d")

    for row in range(1, 6):
        ws1[f"A{row}"].font = Font(bold=True)

    # Summary table by category
    ws1["A7"] = "カテゴリ別CO2排出量サマリー"
    ws1["A7"].font = Font(bold=True, size=12)

    headers = ["カテゴリ", "CO2排出量 (kg-CO2)"]
    for col, h in enumerate(headers, start=1):
        cell = ws1.cell(row=8, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
        cell.border = thin_border

    # Aggregate by category
    category_totals: dict[str, float] = {}
    for r in results_summary:
        cat = r.get("category", "other")
        category_totals[cat] = category_totals.get(cat, 0.0) + r.get("co2_kg", 0.0)

    row_num = 9
    total_co2 = 0.0
    for cat, co2 in sorted(category_totals.items()):
        ws1.cell(row=row_num, column=1, value=cat).border = thin_border
        co2_cell = ws1.cell(row=row_num, column=2, value=round(co2, 3))
        co2_cell.border = thin_border
        co2_cell.alignment = Alignment(horizontal="right")
        co2_cell.number_format = "#,##0.000"
        total_co2 += co2
        row_num += 1

    # Total row
    total_label = ws1.cell(row=row_num, column=1, value="合計")
    total_label.font = Font(bold=True)
    total_label.border = thin_border
    total_value = ws1.cell(row=row_num, column=2, value=round(total_co2, 3))
    total_value.font = Font(bold=True)
    total_value.border = thin_border
    total_value.alignment = Alignment(horizontal="right")
    total_value.number_format = "#,##0.000"

    # Set column widths
    ws1.column_dimensions["A"].width = 25
    ws1.column_dimensions["B"].width = 25

    # --- Sheet 2: 活動量詳細 ---
    ws2 = wb.create_sheet(title="活動量詳細")

    detail_headers = ["カテゴリ", "品目", "数量", "単位", "排出係数 (kg-CO2/unit)", "CO2排出量 (kg-CO2)"]
    for col, h in enumerate(detail_headers, start=1):
        cell = ws2.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
        cell.border = thin_border

    for row_idx, r in enumerate(results_summary, start=2):
        ws2.cell(row=row_idx, column=1, value=r.get("category", "")).border = thin_border
        ws2.cell(row=row_idx, column=2, value=r.get("item_name", "")).border = thin_border
        qty_cell = ws2.cell(row=row_idx, column=3, value=r.get("quantity", 0))
        qty_cell.border = thin_border
        qty_cell.number_format = "#,##0.000"
        ws2.cell(row=row_idx, column=4, value=r.get("unit", "")).border = thin_border
        fv_cell = ws2.cell(row=row_idx, column=5, value=r.get("factor_value", 0))
        fv_cell.border = thin_border
        fv_cell.number_format = "0.000"
        co2_cell = ws2.cell(row=row_idx, column=6, value=round(r.get("co2_kg", 0), 3))
        co2_cell.border = thin_border
        co2_cell.number_format = "#,##0.000"

    # Set column widths
    col_widths = [15, 20, 12, 10, 25, 22]
    for col_idx, width in enumerate(col_widths, start=1):
        ws2.column_dimensions[get_column_letter(col_idx)].width = width

    # --- Sheet 3: Scope別集計 ---
    ws3 = wb.create_sheet(title="Scope別集計")
    scope_rows = scope_summary_from_results(results_summary)
    ws3["A1"] = "Scope別CO2排出量サマリー"
    ws3["A1"].font = Font(bold=True, size=12)
    scope_headers = ["Scope", "CO2排出量 (kg-CO2)", "CO2排出量 (t-CO2)"]
    for col, h in enumerate(scope_headers, start=1):
        cell = ws3.cell(row=3, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border
    for idx, item in enumerate(scope_rows, start=4):
        ws3.cell(row=idx, column=1, value=item["label"]).border = thin_border
        kg_cell = ws3.cell(row=idx, column=2, value=round(item["total_co2_kg"], 3))
        kg_cell.border = thin_border
        kg_cell.number_format = "#,##0.000"
        kg_cell.alignment = Alignment(horizontal="right")
        t_cell = ws3.cell(row=idx, column=3, value=round(item["total_co2_t"], 4))
        t_cell.border = thin_border
        t_cell.number_format = "#,##0.0000"
        t_cell.alignment = Alignment(horizontal="right")
    ws3.column_dimensions["A"].width = 35
    ws3.column_dimensions["B"].width = 22
    ws3.column_dimensions["C"].width = 22

    # Save to bytes
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer.getvalue()


def generate_monthly_report_csv(project, month: str, results_summary: list[dict]) -> bytes:
    """Generate a UTF-8 BOM CSV report (Excel-compatible)."""
    buffer = io.StringIO()
    writer = csv.writer(buffer, lineterminator="\r\n")
    writer.writerow(["プロジェクト名", getattr(project, "name", "")])
    writer.writerow(["支店", getattr(project, "branch", "")])
    writer.writerow(["工種", getattr(project, "work_type", "")])
    writer.writerow(["対象月", month])
    writer.writerow(["レポート作成日", datetime.date.today().strftime("%Y-%m-%d")])
    writer.writerow([])
    writer.writerow(["=== Scope別集計 ==="])
    for item in scope_summary_from_results(results_summary):
        writer.writerow([item["label"], round(item["total_co2_kg"], 3), round(item["total_co2_t"], 4)])
    writer.writerow([])
    writer.writerow(["カテゴリ", "品目", "数量", "単位", "排出係数 (kg-CO2/unit)", "CO2排出量 (kg-CO2)"])
    for r in results_summary:
        writer.writerow([
            r.get("category", ""),
            r.get("item_name", ""),
            r.get("quantity", 0),
            r.get("unit", ""),
            r.get("factor_value", 0),
            round(r.get("co2_kg", 0), 3),
        ])
    total = sum(r.get("co2_kg", 0) for r in results_summary)
    writer.writerow(["合計", "", "", "", "", round(total, 3)])
    return ("\ufeff" + buffer.getvalue()).encode("utf-8")


def generate_activity_import_template() -> bytes:
    """Generate an Excel template for bulk activity import."""
    wb = Workbook()
    ws = wb.active
    ws.title = "活動量"
    headers = ["project_id", "target_month", "category", "item_name", "quantity", "unit", "source_file", "note"]
    for col, h in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="2E75B6", end_color="2E75B6", fill_type="solid")
        cell.font = Font(bold=True, color="FFFFFF")
    examples = [
        ["<プロジェクトID>", "2026-08", "fuel", "軽油", 500.0, "L", "給油伝票-202608-001", "現場タンク給油"],
        ["<プロジェクトID>", "2026-08", "power", "電力", 1200.0, "kWh", "電力検針票", ""],
        ["<プロジェクトID>", "2026-08", "material", "生コン", 12.5, "t", "納品書", ""],
        ["<プロジェクトID>", "2026-08", "transport", "一般輸送", 320.0, "t-km", "運送依頼書", ""],
    ]
    for row in examples:
        ws.append(row)
    for col_idx, width in enumerate([22, 12, 14, 22, 12, 10, 26, 18], start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = width
    ws2 = wb.create_sheet(title="カテゴリ一覧")
    ws2.append(["カテゴリ", "説明"])
    for cat, desc in [
        ("fuel", "燃料 (軽油・A重油・ガソリン等)"),
        ("power", "電力 (kWh)"),
        ("material", "材料 (t/kg)"),
        ("transport", "輸送 (t-km)"),
        ("machine", "建機稼働"),
        ("ship", "船舶稼働"),
        ("waste", "廃棄物"),
    ]:
        ws2.append([cat, desc])
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer.getvalue()


def generate_project_import_template() -> bytes:
    """Generate an Excel template for bulk project import."""
    wb = Workbook()
    ws = wb.active
    ws.title = "工事"
    headers = ["name", "branch", "work_type", "start_date", "end_date", "description", "close_day"]
    for col, h in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="2E75B6", end_color="2E75B6", fill_type="solid")
    ws.append(["○○道路改良工事", "東京支店", "道路工事", "2026-04-01", "2026-09-30", "PoC用", 25])
    ws.append(["○○港湾護岸工事", "大阪支店", "港湾工事", "2026-01-01", "2026-12-31", "PoC用", 25])
    for col_idx, width in enumerate([28, 14, 14, 14, 14, 24, 10], start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = width
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer.getvalue()


def generate_monthly_report_pdf(project, month: str, results_summary: list[dict]) -> bytes:
    """Generate a Japanese-friendly PDF monthly report (A4)."""
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(
        buffer,
        pagesize=A4,
        rightMargin=18 * mm,
        leftMargin=18 * mm,
        topMargin=18 * mm,
        bottomMargin=18 * mm,
        title=f"CO2排出量月次レポート {month}",
        author="MIRAI Site Carbon Navigator",
    )

    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        "TitleJP",
        parent=styles["Title"],
        fontName=_JAPANESE_FONT,
        fontSize=16,
        leading=22,
        textColor=colors.HexColor("#1f5e33"),
        spaceAfter=6,
    )
    h2_style = ParagraphStyle(
        "H2JP",
        parent=styles["Heading2"],
        fontName=_JAPANESE_FONT,
        fontSize=12,
        leading=16,
        textColor=colors.HexColor("#2d7d46"),
        spaceBefore=10,
        spaceAfter=6,
    )
    body_style = ParagraphStyle(
        "BodyJP",
        parent=styles["BodyText"],
        fontName=_JAPANESE_FONT,
        fontSize=9,
        leading=13,
    )
    elements = []
    elements.append(Paragraph("CO2排出量 月次レポート", title_style))
    elements.append(Paragraph(f"対象月: {month} / 作成日: {datetime.date.today().isoformat()}", body_style))
    elements.append(Spacer(1, 6 * mm))

    info = Table(
        [
            ["プロジェクト名", getattr(project, "name", "")],
            ["支店", getattr(project, "branch", "")],
            ["工種", getattr(project, "work_type", "")],
            ["期間", f"{getattr(project, 'start_date', '')} 〜 {getattr(project, 'end_date', '')}"],
        ],
        colWidths=[35 * mm, 139 * mm],
    )
    info.setStyle(TableStyle([
        ("FONTNAME", (0, 0), (-1, -1), _JAPANESE_FONT),
        ("FONTSIZE", (0, 0), (-1, -1), 9),
        ("BACKGROUND", (0, 0), (0, -1), colors.HexColor("#e8f5ec")),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#c8e6d0")),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("TOPPADDING", (0, 0), (-1, -1), 4),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
    ]))
    elements.append(info)
    elements.append(Spacer(1, 4 * mm))

    elements.append(Paragraph("1. カテゴリ別CO2排出量サマリー", h2_style))
    category_totals: dict[str, float] = {}
    for r in results_summary:
        cat = r.get("category", "other")
        category_totals[cat] = category_totals.get(cat, 0.0) + r.get("co2_kg", 0.0)
    cat_rows = [["カテゴリ", "CO2排出量 (kg-CO2)", "CO2排出量 (t-CO2)"]]
    for cat, co2 in sorted(category_totals.items()):
        cat_rows.append([cat, f"{co2:,.3f}", f"{co2 / 1000:,.4f}"])
    total_co2 = sum(category_totals.values())
    cat_rows.append(["合計", f"{total_co2:,.3f}", f"{total_co2 / 1000:,.4f}"])
    cat_table = Table(cat_rows, colWidths=[70 * mm, 52 * mm, 52 * mm])
    cat_table.setStyle(TableStyle([
        ("FONTNAME", (0, 0), (-1, -1), _JAPANESE_FONT),
        ("FONTSIZE", (0, 0), (-1, -1), 9),
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#2E75B6")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
        ("ALIGN", (1, 0), (-1, -1), "RIGHT"),
        ("FONTNAME", (0, -1), (-1, -1), _JAPANESE_FONT),
        ("BACKGROUND", (0, -1), (-1, -1), colors.HexColor("#e8f5ec")),
    ]))
    elements.append(cat_table)

    elements.append(Paragraph("2. 活動量詳細", h2_style))
    detail_rows = [["カテゴリ", "品目", "数量", "単位", "排出係数", "CO2 (kg-CO2)"]]
    for r in results_summary:
        detail_rows.append([
            r.get("category", ""),
            r.get("item_name", ""),
            f"{r.get('quantity', 0):,.3f}",
            r.get("unit", ""),
            f"{r.get('factor_value', 0):,.4f}",
            f"{r.get('co2_kg', 0):,.3f}",
        ])
    detail_table = Table(detail_rows, colWidths=[28 * mm, 42 * mm, 26 * mm, 18 * mm, 30 * mm, 30 * mm])
    detail_table.setStyle(TableStyle([
        ("FONTNAME", (0, 0), (-1, -1), _JAPANESE_FONT),
        ("FONTSIZE", (0, 0), (-1, -1), 8),
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#2E75B6")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("GRID", (0, 0), (-1, -1), 0.4, colors.grey),
        ("ALIGN", (2, 0), (-1, -1), "RIGHT"),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("TOPPADDING", (0, 0), (-1, -1), 3),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 3),
    ]))
    elements.append(detail_table)

    elements.append(Paragraph("Scope別集計", h2_style))
    scope_rows = [["Scope", "CO2排出量 (kg-CO2)", "CO2排出量 (t-CO2)"]]
    for item in scope_summary_from_results(results_summary):
        scope_rows.append([item["label"], f"{item['total_co2_kg']:,.3f}", f"{item['total_co2_t']:,.4f}"])
    scope_table = Table(scope_rows, colWidths=[80 * mm, 47 * mm, 47 * mm])
    scope_table.setStyle(TableStyle([
        ("FONTNAME", (0, 0), (-1, -1), _JAPANESE_FONT),
        ("FONTSIZE", (0, 0), (-1, -1), 9),
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#2E75B6")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
        ("ALIGN", (1, 0), (-1, -1), "RIGHT"),
    ]))
    elements.append(scope_table)

    suggestions = get_reduction_suggestions(
        [{"category": r.get("category"), "co2_kg": r.get("co2_kg")} for r in results_summary]
    )
    if suggestions:
        elements.append(Paragraph("3. 削減ナビ（推奨アクション）", h2_style))
        for s in suggestions[:3]:
            text = f"<b>{s['category']}（{s['total_co2_kg'] / 1000:,.3f} t-CO2）</b><br/>" + "<br/>".join(
                f"・{item}" for item in s["suggestions"][:3]
            )
            elements.append(Paragraph(text, body_style))
            elements.append(Spacer(1, 2 * mm))

    elements.append(Spacer(1, 4 * mm))
    elements.append(
        Paragraph(
            "※ 排出係数は対象月時点で有効な最新版を適用しています。算定根拠（係数値・出典・適用開始日）はシステム上で追跡可能です。",
            body_style,
        )
    )

    doc.build(elements)
    buffer.seek(0)
    return buffer.getvalue()


def generate_project_card_pdf(
    project,
    monthly_trend: list[dict],
    category_totals: dict[str, float],
    reduction_actions: list[dict],
    feedbacks: list[dict],
) -> bytes:
    """Generate a project dossier PDF (工事カルテ) summarizing the whole project."""
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(
        buffer,
        pagesize=A4,
        rightMargin=18 * mm,
        leftMargin=18 * mm,
        topMargin=18 * mm,
        bottomMargin=18 * mm,
        title=f"工事カルテ {getattr(project, 'name', '')}",
        author="MIRAI Site Carbon Navigator",
    )
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        "TitleJP",
        parent=styles["Title"],
        fontName=_JAPANESE_FONT,
        fontSize=16,
        leading=22,
        textColor=colors.HexColor("#1f5e33"),
        spaceAfter=6,
    )
    h2_style = ParagraphStyle(
        "H2JP",
        parent=styles["Heading2"],
        fontName=_JAPANESE_FONT,
        fontSize=12,
        leading=16,
        textColor=colors.HexColor("#2d7d46"),
        spaceBefore=10,
        spaceAfter=6,
    )
    body_style = ParagraphStyle(
        "BodyJP",
        parent=styles["BodyText"],
        fontName=_JAPANESE_FONT,
        fontSize=9,
        leading=13,
    )

    elements = []
    elements.append(Paragraph("工事カルテ（CO2排出サマリー）", title_style))
    elements.append(Paragraph(f"作成日: {datetime.date.today().isoformat()}", body_style))
    elements.append(Spacer(1, 4 * mm))

    info = Table(
        [
            ["プロジェクト名", getattr(project, "name", "")],
            ["支店", getattr(project, "branch", "")],
            ["工種", getattr(project, "work_type", "")],
            ["期間", f"{getattr(project, 'start_date', '')} 〜 {getattr(project, 'end_date', '')}"],
        ],
        colWidths=[35 * mm, 139 * mm],
    )
    info.setStyle(TableStyle([
        ("FONTNAME", (0, 0), (-1, -1), _JAPANESE_FONT),
        ("FONTSIZE", (0, 0), (-1, -1), 9),
        ("BACKGROUND", (0, 0), (0, -1), colors.HexColor("#e8f5ec")),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#c8e6d0")),
    ]))
    elements.append(info)
    elements.append(Spacer(1, 4 * mm))

    total_kg = sum(category_totals.values())
    elements.append(Paragraph(f"1. 排出量サマリー（全期間合計 {total_kg / 1000:,.3f} t-CO2）", h2_style))
    cat_rows = [["カテゴリ", "CO2排出量 (kg-CO2)", "CO2排出量 (t-CO2)", "割合"]]
    for cat, co2 in sorted(category_totals.items(), key=lambda x: -x[1]):
        ratio = f"{co2 / total_kg * 100:.1f}%" if total_kg else "-"
        cat_rows.append([cat, f"{co2:,.3f}", f"{co2 / 1000:,.4f}", ratio])
    cat_table = Table(cat_rows, colWidths=[55 * mm, 45 * mm, 45 * mm, 29 * mm])
    cat_table.setStyle(TableStyle([
        ("FONTNAME", (0, 0), (-1, -1), _JAPANESE_FONT),
        ("FONTSIZE", (0, 0), (-1, -1), 9),
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#2E75B6")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
        ("ALIGN", (1, 0), (-1, -1), "RIGHT"),
    ]))
    elements.append(cat_table)

    elements.append(Paragraph("2. 月次排出推移", h2_style))
    trend_rows = [["対象月", "CO2排出量 (kg-CO2)", "CO2排出量 (t-CO2)"]]
    for item in monthly_trend:
        trend_rows.append([
            item["target_month"],
            f"{item['total_co2_kg']:,.3f}",
            f"{item['total_co2_kg'] / 1000:,.4f}",
        ])
    trend_table = Table(trend_rows, colWidths=[45 * mm, 65 * mm, 65 * mm])
    trend_table.setStyle(TableStyle([
        ("FONTNAME", (0, 0), (-1, -1), _JAPANESE_FONT),
        ("FONTSIZE", (0, 0), (-1, -1), 9),
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#2E75B6")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
        ("ALIGN", (1, 0), (-1, -1), "RIGHT"),
    ]))
    elements.append(trend_table)

    if reduction_actions:
        elements.append(Paragraph("3. 削減アクション", h2_style))
        status_labels = {"planned": "計画", "implemented": "実施済み", "declined": "見送り"}
        action_rows = [["対象月", "カテゴリ", "提案", "状態", "想定(kg)", "実績(kg)"]]
        for a in reduction_actions[:10]:
            action_rows.append([
                a["target_month"],
                a["category"],
                a["suggestion"],
                status_labels.get(a["status"], a["status"]),
                f"{a.get('estimated_reduction_kg') or 0:,.1f}",
                f"{a.get('actual_reduction_kg') or 0:,.1f}",
            ])
        action_table = Table(action_rows, colWidths=[22 * mm, 20 * mm, 60 * mm, 22 * mm, 25 * mm, 25 * mm])
        action_table.setStyle(TableStyle([
            ("FONTNAME", (0, 0), (-1, -1), _JAPANESE_FONT),
            ("FONTSIZE", (0, 0), (-1, -1), 8),
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#2E75B6")),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("GRID", (0, 0), (-1, -1), 0.4, colors.grey),
        ]))
        elements.append(action_table)

    if feedbacks:
        elements.append(Paragraph("4. 現場フィードバック", h2_style))
        for f in feedbacks[:5]:
            elements.append(Paragraph(f"・[{f['target_month']}] {f['content']}", body_style))

    doc.build(elements)
    buffer.seek(0)
    return buffer.getvalue()


def generate_annual_report_pdf(
    year: int,
    projects_summary: list[dict],
    scope_totals: dict[str, float],
    sbti_progress: list[dict],
    actions_summary: dict,
    credits_summary: dict,
) -> bytes:
    """Generate an annual environmental report PDF (年次環境報告書)."""
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(
        buffer,
        pagesize=A4,
        rightMargin=18 * mm,
        leftMargin=18 * mm,
        topMargin=18 * mm,
        bottomMargin=18 * mm,
        title=f"年次環境報告書 {year}",
        author="MIRAI Site Carbon Navigator",
    )
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        "TitleJP", parent=styles["Title"], fontName=_JAPANESE_FONT,
        fontSize=18, leading=24, textColor=colors.HexColor("#1f5e33"), spaceAfter=4,
    )
    h2_style = ParagraphStyle(
        "H2JP", parent=styles["Heading2"], fontName=_JAPANESE_FONT,
        fontSize=13, leading=17, textColor=colors.HexColor("#2d7d46"),
        spaceBefore=12, spaceAfter=6,
    )
    body_style = ParagraphStyle(
        "BodyJP", parent=styles["BodyText"], fontName=_JAPANESE_FONT,
        fontSize=9, leading=13,
    )
    _table_style = [
        ("FONTNAME", (0, 0), (-1, -1), _JAPANESE_FONT),
        ("FONTSIZE", (0, 0), (-1, -1), 8),
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#2E75B6")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("GRID", (0, 0), (-1, -1), 0.4, colors.grey),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
    ]

    elements = []
    elements.append(Paragraph(f"年次環境報告書 {year}", title_style))
    elements.append(Paragraph(
        f"作成日: {datetime.date.today().isoformat()}／MIRAI Site Carbon Navigator",
        body_style,
    ))
    elements.append(Spacer(1, 6 * mm))

    total_kg = sum(scope_totals.values())
    elements.append(Paragraph(f"1. 全体サマリー（総排出量 {total_kg / 1000:,.3f} t-CO2）", h2_style))
    overview_rows = [["Scope", "CO2排出量 (kg-CO2)", "CO2排出量 (t-CO2)"]]
    for scope, label in [("scope1", "Scope1 直接排出"), ("scope2", "Scope2 エネルギー間接排出"), ("scope3", "Scope3 その他間接排出")]:
        kg = scope_totals.get(scope, 0.0)
        overview_rows.append([label, f"{kg:,.3f}", f"{kg / 1000:,.4f}"])
    overview_table = Table(overview_rows, colWidths=[80 * mm, 47 * mm, 47 * mm])
    overview_table.setStyle(TableStyle(_table_style))
    elements.append(overview_table)

    elements.append(Paragraph("2. 工事別排出量", h2_style))
    project_rows = [["工事名", "支店", "工種", "CO2排出量 (t-CO2)"]]
    for p in sorted(projects_summary, key=lambda x: -x["total_co2_kg"]):
        project_rows.append([
            p["name"], p["branch"] or "-", p["work_type"] or "-",
            f"{p['total_co2_kg'] / 1000:,.4f}",
        ])
    project_table = Table(project_rows, colWidths=[70 * mm, 30 * mm, 35 * mm, 39 * mm])
    project_table.setStyle(TableStyle(_table_style))
    elements.append(project_table)

    if sbti_progress:
        elements.append(Paragraph("3. SBTi目標進捗", h2_style))
        sbti_rows = [["目標", "Scope", "基準(t)", "目標(t)", "現状(t)", "達成率", "判定"]]
        for item in sbti_progress:
            sbti_rows.append([
                item["name"],
                item["scope"],
                f"{item['base_emissions_kg'] / 1000:,.2f}",
                f"{item['target_emissions_kg'] / 1000:,.2f}",
                f"{item['current_emissions_kg'] / 1000:,.2f}",
                f"{max(0.0, item['reduction_achieved_percent']):.1f}%",
                "順調" if item["on_track"] else "遅延",
            ])
        sbti_table = Table(sbti_rows, colWidths=[45 * mm, 18 * mm, 25 * mm, 25 * mm, 25 * mm, 20 * mm, 16 * mm])
        sbti_table.setStyle(TableStyle(_table_style))
        elements.append(sbti_table)

    elements.append(Paragraph("4. 削減アクション実績", h2_style))
    actions_text = (
        f"実施済み {actions_summary.get('implemented', 0)} 件／"
        f"実績削減量合計 {actions_summary.get('total_reduction_kg', 0.0) / 1000:,.3f} t-CO2"
    )
    elements.append(Paragraph(actions_text, body_style))

    elements.append(Paragraph("5. カーボンクレジット（オフセット）", h2_style))
    credits_text = (
        f"保有合計 {credits_summary.get('total_tco2', 0.0):,.3f} t-CO2"
        f"（利用可能 {credits_summary.get('available_tco2', 0.0):,.3f} / "
        f"充当済み {credits_summary.get('allocated_tco2', 0.0):,.3f} / "
        f"無効化 {credits_summary.get('retired_tco2', 0.0):,.3f}）"
    )
    elements.append(Paragraph(credits_text, body_style))

    doc.build(elements)
    buffer.seek(0)
    return buffer.getvalue()
